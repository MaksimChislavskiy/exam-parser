from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font

from .models import TaskRecord, VariantMetadata


BASE_HEADERS = ("task_num", "condition", "image_name", "solution", "answer")
CLASSIFICATION_HEADERS = ("exams_id", "topics_id")
HEADERS = BASE_HEADERS + CLASSIFICATION_HEADERS
ABOUT_HEADERS = (
    "class",
    "year",
    "date",
    "topic",
    "exam_id",
    "title",
    "code",
    "source_name",
    "is_public",
    "source_url",
    "description",
)
HTML_BLOCK_PATTERN = re.compile(
    r"</?(?:p|div|figure|center|ul|ol|li|table|thead|tbody|tr|td|th)\b",
    re.IGNORECASE,
)
SUBPART_MARKER_PATTERN = re.compile(
    r"(?<!\S)[a-eа-е]\)\s+",
    re.IGNORECASE,
)


def read_tasks_xlsx(input_path: str | Path) -> list[TaskRecord]:
    input_path = Path(input_path)
    workbook = load_workbook(input_path, read_only=True, data_only=False)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = tuple(next(rows, ()))
        if headers not in {BASE_HEADERS, HEADERS}:
            raise ValueError(
                f"Некорректные столбцы в {input_path}: "
                f"ожидались {BASE_HEADERS} или {HEADERS}, получены {headers}"
            )

        records: list[TaskRecord] = []
        task_numbers: set[str] = set()
        for row_number, row in enumerate(rows, start=2):
            values = tuple(row[: len(headers)])
            values += (None,) * (len(headers) - len(values))
            if not any(_cell_text(value).strip() for value in values):
                continue

            task_num = _cell_text(values[0]).strip()
            condition = _cell_text(values[1])
            if not task_num:
                raise ValueError(
                    f"В строке {row_number} файла {input_path} отсутствует task_num"
                )
            if not condition.strip():
                raise ValueError(
                    f"В строке {row_number} файла {input_path} отсутствует condition"
                )
            if task_num in task_numbers:
                raise ValueError(
                    f"В файле {input_path} повторяется задача {task_num}"
                )
            task_numbers.add(task_num)

            image_name = _cell_text(values[2]).strip() or None
            records.append(
                TaskRecord(
                    task_num=task_num,
                    condition=condition,
                    image_name=image_name,
                    solution=_cell_text(values[3]),
                    answer=_cell_text(values[4]),
                    exams_id=(
                        _optional_int(
                            values[5],
                            column="exams_id",
                            row_number=row_number,
                            input_path=input_path,
                        )
                        if len(headers) > 5
                        else None
                    ),
                    topics_id=(
                        _optional_int(
                            values[6],
                            column="topics_id",
                            row_number=row_number,
                            input_path=input_path,
                        )
                        if len(headers) > 6
                        else None
                    ),
                )
            )

        if not records:
            raise ValueError(f"В файле {input_path} нет задач для продолжения")
        return records
    finally:
        workbook.close()


def read_variant_metadata_xlsx(
    input_path: str | Path,
) -> VariantMetadata | None:
    input_path = Path(input_path)
    workbook = load_workbook(input_path, read_only=True, data_only=False)
    try:
        if "about" not in workbook.sheetnames:
            return None

        sheet = workbook["about"]
        rows = sheet.iter_rows(min_row=1, max_row=2, values_only=True)
        headers = tuple(next(rows, ()))
        if headers != ABOUT_HEADERS:
            raise ValueError(
                f"Некорректные столбцы листа about в {input_path}: "
                f"ожидались {ABOUT_HEADERS}, получены {headers}"
            )

        values = tuple(next(rows, ()))
        values += (None,) * (len(ABOUT_HEADERS) - len(values))
        if not any(_cell_text(value).strip() for value in values):
            return None

        payload = dict(zip(ABOUT_HEADERS, values[: len(ABOUT_HEADERS)]))
        return VariantMetadata.model_validate(payload)
    finally:
        workbook.close()


def write_tasks_xlsx(
    records: list[TaskRecord],
    output_path: str | Path,
    *,
    about: VariantMetadata | None = None,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # При checkpoint/resume не теряем уже заполненный лист about.
    if about is None and output_path.is_file():
        try:
            about = read_variant_metadata_xlsx(output_path)
        except ValueError:
            about = None

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tasks"
    sheet.append(HEADERS)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for record in records:
        sheet.append(
            (
                record.task_num,
                _excel_safe(_condition_html(record.condition)),
                record.image_name,
                _excel_safe(record.solution),
                _excel_safe(record.answer),
                record.exams_id,
                record.topics_id,
            )
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in {
        "A": 14,
        "B": 70,
        "C": 28,
        "D": 90,
        "E": 24,
        "F": 14,
        "G": 14,
    }.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    about_sheet = workbook.create_sheet("about")
    about_sheet.append(ABOUT_HEADERS)
    for cell in about_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if about is None:
        about_sheet.append((None,) * len(ABOUT_HEADERS))
    else:
        payload = about.model_dump(by_alias=True)
        about_sheet.append(tuple(payload[header] for header in ABOUT_HEADERS))

    for column, width in {
        "A": 10,
        "B": 10,
        "C": 14,
        "D": 10,
        "E": 10,
        "F": 38,
        "G": 18,
        "H": 20,
        "I": 12,
        "J": 40,
        "K": 80,
    }.items():
        about_sheet.column_dimensions[column].width = width
    for row in about_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(output_path)
    return output_path


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _optional_int(
    value: object,
    *,
    column: str,
    row_number: int,
    input_path: Path,
) -> int | None:
    text = _cell_text(value).strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError as error:
        raise ValueError(
            f"В строке {row_number} файла {input_path} "
            f"некорректный {column}: {text!r}"
        ) from error


def _excel_safe(value: str) -> str:
    return ILLEGAL_CHARACTERS_RE.sub("", value)


def _condition_html(value: str) -> str:
    """Оборачивает только подпункты условия в HTML-абзацы для импорта на сайт."""
    normalized = value.replace("\r\n", "\n").replace("\r", "\n").strip()
    if not normalized or HTML_BLOCK_PATTERN.search(normalized):
        return normalized

    split = _split_subparts(normalized)
    if split is None:
        return normalized

    prefix, subparts = split
    tagged_subparts = [f"<p>{subpart}</p>" for subpart in subparts]
    if prefix:
        return "\n".join((prefix, *tagged_subparts))
    return "\n".join(tagged_subparts)


def _split_subparts(value: str) -> tuple[str, list[str]] | None:
    markers = list(SUBPART_MARKER_PATTERN.finditer(value))
    if len(markers) < 2:
        return None

    prefix = value[: markers[0].start()].strip()
    subparts: list[str] = []
    for index, marker in enumerate(markers):
        end = markers[index + 1].start() if index + 1 < len(markers) else len(value)
        subpart = value[marker.start() : end].strip()
        subparts.append(" ".join(line.strip() for line in subpart.splitlines()))
    return prefix, subparts
