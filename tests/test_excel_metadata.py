from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from exam_parser.excel import (
    ABOUT_HEADERS,
    BASE_HEADERS,
    HEADERS,
    read_tasks_xlsx,
    read_variant_metadata_xlsx,
    write_tasks_xlsx,
)
from exam_parser.models import TaskRecord, VariantMetadata


def test_writer_adds_classification_columns_and_about_sheet(tmp_path: Path) -> None:
    output_path = tmp_path / "tasks.xlsx"
    write_tasks_xlsx(
        [
            TaskRecord(
                task_num="1",
                condition="Найдите угол.",
                image_name="task_1.png",
                exams_id=171,
                topics_id=163,
            )
        ],
        output_path,
        about=VariantMetadata.model_validate(
            {
                "class": 10,
                "year": 2026,
                "date": "2026-02-11",
                "topic": 1,
                "exam_id": 2,
                "title": "Тренировочная работа №1 по математике",
                "code": "МА2500109",
                "source_name": "СтатГрад",
                "is_public": True,
            }
        ),
    )

    workbook = load_workbook(output_path, read_only=True, data_only=False)
    try:
        tasks = workbook["Tasks"]
        assert tuple(cell.value for cell in tasks[1]) == HEADERS
        assert tasks["F2"].value == 171
        assert tasks["G2"].value == 163

        about = workbook["about"]
        assert tuple(cell.value for cell in about[1]) == ABOUT_HEADERS
        assert about["A2"].value == 10
        assert about["G2"].value == "МА2500109"
    finally:
        workbook.close()


def test_reader_accepts_legacy_five_column_checkpoint(tmp_path: Path) -> None:
    input_path = tmp_path / "legacy.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(BASE_HEADERS)
    sheet.append(("1", "Условие", None, "", "42"))
    workbook.save(input_path)

    records = read_tasks_xlsx(input_path)

    assert len(records) == 1
    assert records[0].answer == "42"
    assert records[0].exams_id is None
    assert records[0].topics_id is None


def test_checkpoint_rewrite_preserves_existing_about_sheet(tmp_path: Path) -> None:
    output_path = tmp_path / "tasks.xlsx"
    metadata = VariantMetadata.model_validate(
        {
            "class": 10,
            "year": 2026,
            "code": "МА2500109",
            "source_name": "СтатГрад",
        }
    )
    write_tasks_xlsx(
        [TaskRecord(task_num="1", condition="Первое условие")],
        output_path,
        about=metadata,
    )

    write_tasks_xlsx(
        [TaskRecord(task_num="1", condition="Исправленное условие")],
        output_path,
    )

    preserved = read_variant_metadata_xlsx(output_path)
    assert preserved is not None
    assert preserved.school_class == 10
    assert preserved.code == "МА2500109"
    assert preserved.source_name == "СтатГрад"
